"""
etl/load_headstart_pir.py -- Load Head Start PIR (Program Information Report) data.

The PIR is the annual census of all Head Start and Early Head Start programs.
It contains program-level data on enrollment, capacity, staffing, health services,
and demographics -- everything needed to evaluate community facility investment
opportunities in the early childhood space.

Data source:
    HSES (Head Start Enterprise System) at https://hses.ohs.acf.hhs.gov
    Requires an account + Data Use Agreement. Federal public-domain data
    behind a procedural gate.

    Export path: HSES > Reports > PIR Data > Data as of Date Report
    File format: .xlsx or .xls with sheets: Section A, B, C, D, Program Details,
    Reference, Configuration.

    Key structural quirk: Sections A-D have TWO header rows.
      Row 1 = descriptive column names (long text)
      Row 2 = PIR question codes (A.1.a, B.3-1, etc.)
    Program Details has one header row with normal column names.

    Question codes shifted across years, so we map by DESCRIPTION text (more
    stable) with fallback to known codes per era.

Usage:
    # Single file
    python etl/load_headstart_pir.py --file data/raw/childcare/PIR_Export_2025.xlsx

    # Batch load all files in a directory
    python etl/load_headstart_pir.py --dir data/raw/childcare

    # Filter
    python etl/load_headstart_pir.py --dir data/raw/childcare --states GA TX
    python etl/load_headstart_pir.py --file data/raw/childcare/PIR_Export_2025.xlsx --dry-run
    python etl/load_headstart_pir.py --file data/raw/childcare/PIR_Export_2025.xlsx --columns-only
"""

import argparse
import os
import re
import sys

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
import db


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def to_int(v):
    if v is None or v == "" or v == " ":
        return None
    try:
        return int(float(v))
    except (ValueError, TypeError):
        return None


def to_str(v):
    if v is None:
        return None
    s = str(v).strip()
    return s if s else None


def infer_year(filename):
    m = re.search(r'(\d{4})', os.path.basename(filename))
    return int(m.group(1)) if m else None


# ---------------------------------------------------------------------------
# Description-based field finder
#
# PIR question codes shift between years, but the descriptive text is more
# stable. We search for keywords in the description row to find the right
# column index, then read the code from the code row. This makes the loader
# resilient across 2008-2025+.
# ---------------------------------------------------------------------------

def _find_col(descs, codes, *keywords):
    """
    Find column index where ALL keywords appear in the description (case-insensitive).
    Returns the question code at that index, or None.
    """
    for i, desc in enumerate(descs):
        if desc is None:
            continue
        d = str(desc).lower()
        if all(kw.lower() in d for kw in keywords):
            return codes[i] if i < len(codes) else None
    return None


def build_code_map(descs, codes):
    """
    Build a mapping from our canonical field names to actual question codes
    for this particular file. Uses description-keyword matching with fallbacks.
    """
    m = {}

    # Section A fields
    m["funded_enrollment"] = (
        _find_col(descs, codes, "ACF", "Funded", "Enrollment")
    )
    m["non_acf_enrollment"] = (
        _find_col(descs, codes, "Non ACF", "Funded", "Enrollment")
        or _find_col(descs, codes, "Non-ACF", "Enrollment")
    )
    m["total_cumulative_enrollment"] = (
        _find_col(descs, codes, "Total cumulative enrollment")
        or _find_col(descs, codes, "Total Cumulative Enrollment")
    )
    m["total_slots_center_based"] = (
        _find_col(descs, codes, "Total number of slots", "center-based")
        or _find_col(descs, codes, "Total Funded Enrollment")  # older years
    )
    m["slots_at_child_care_partner"] = (
        _find_col(descs, codes, "slots at a child care partner")
        or _find_col(descs, codes, "Funded Enrollment at Center-based Child Care Partner")
    )
    m["total_classes"] = (
        _find_col(descs, codes, "Total Classes Operated")
    )
    m["home_based_slots"] = (
        _find_col(descs, codes, "Home-based")
    )
    m["family_child_care_slots"] = (
        _find_col(descs, codes, "Family Child Care Option")
    )

    # Age breakdown
    m["children_lt1"] = _find_col(descs, codes, "Less than 1 Year")
    m["children_1yr"] = _find_col(descs, codes, "1 Year Old")
    m["children_2yr"] = _find_col(descs, codes, "2 Years Old")
    m["children_3yr"] = _find_col(descs, codes, "3 Years Old")
    m["children_4yr"] = _find_col(descs, codes, "4 Years Old")
    m["children_5plus"] = _find_col(descs, codes, "5 Years and Older")
    m["pregnant_women"] = _find_col(descs, codes, "Pregnant Women")

    # Eligibility
    m["eligible_income"] = (
        _find_col(descs, codes, "100%", "federal")
        or _find_col(descs, codes, "Income Eligibility")
        or _find_col(descs, codes, "Income at or below")
    )
    m["eligible_public_assist"] = (
        _find_col(descs, codes, "Public assistance")
        or _find_col(descs, codes, "Receipt of Public Assistance")
    )
    m["eligible_foster"] = (
        _find_col(descs, codes, "Foster")
    )
    m["eligible_homeless"] = (
        _find_col(descs, codes, "Homeless")
    )

    # Turnover
    m["children_left_program"] = (
        _find_col(descs, codes, "Preschool children who left")
        or _find_col(descs, codes, "children who left the program")
    )
    m["children_end_of_year"] = (
        _find_col(descs, codes, "enrolled in Head Start at the end")
        or _find_col(descs, codes, "enrolled at the end of the")
    )

    # Demographics
    m["dual_language_learners"] = (
        _find_col(descs, codes, "Dual Language Learners")
    )
    m["children_transported"] = (
        _find_col(descs, codes, "Children Transported")
        or _find_col(descs, codes, "Number of Children Transported")
    )
    m["children_with_subsidy"] = (
        _find_col(descs, codes, "Child Care Subsidy")
        or _find_col(descs, codes, "Receiving Child Care Subsidy")
    )

    return {k: v for k, v in m.items() if v is not None}


def build_staff_code_map(descs, codes):
    """Build code map for Section B (staffing)."""
    m = {}
    m["total_staff"] = (
        _find_col(descs, codes, "Total Head Start Staff")
        or _find_col(descs, codes, "Total Staff")
    )
    m["total_contracted_staff"] = (
        _find_col(descs, codes, "Total Contracted Staff")
    )
    m["classroom_teachers"] = (
        _find_col(descs, codes, "Classroom Teachers")
    )
    m["assistant_teachers"] = (
        _find_col(descs, codes, "Assistant Teachers")
    )
    m["volunteers"] = (
        _find_col(descs, codes, "Total Volunteers")
    )
    # BA+ teachers: advanced degree + baccalaureate lines
    m["_teachers_advanced"] = (
        _find_col(descs, codes, "Advanced degree", "early childhood", "Classroom Teachers")
        or _find_col(descs, codes, "advanced degree", "Classroom Teachers")
    )
    m["_teachers_bachelors"] = (
        _find_col(descs, codes, "baccalaureate degree", "early childhood", "Classroom Teachers")
        or _find_col(descs, codes, "baccalaureate degree", "Classroom Teachers")
    )
    return {k: v for k, v in m.items() if v is not None}


def build_health_code_map(descs, codes):
    """Build code map for Section C (health)."""
    m = {}
    m["children_with_insurance_start"] = (
        _find_col(descs, codes, "Health Insurance", "Enrollment")
        or _find_col(descs, codes, "health insurance at enrollment")
    )
    m["children_with_insurance_end"] = (
        _find_col(descs, codes, "Health Insurance", "End")
    )
    m["children_medicaid_start"] = (
        _find_col(descs, codes, "Medicaid", "Enrollment")
    )
    m["children_no_insurance_start"] = (
        _find_col(descs, codes, "no health insurance at enrollment")
        or _find_col(descs, codes, "no health insurance")
    )
    m["children_with_medical_home_start"] = (
        _find_col(descs, codes, "ongoing source of continuous")
        or _find_col(descs, codes, "medical home")
    )
    m["children_at_fqhc_start"] = (
        _find_col(descs, codes, "federally qualified Health Center")
        or _find_col(descs, codes, "FQHC")
    )
    return {k: v for k, v in m.items() if v is not None}


def build_admin_code_map(descs, codes):
    """Build code map for Section D (admin)."""
    m = {}
    m["child_care_partners"] = (
        _find_col(descs, codes, "child care partners", "formal agreement")
        or _find_col(descs, codes, "child care partners")
    )
    m["leas_in_service_area"] = (
        _find_col(descs, codes, "LEAs in the service area")
        or _find_col(descs, codes, "total number of LEAs")
    )
    return {k: v for k, v in m.items() if v is not None}


# ---------------------------------------------------------------------------
# File readers (xlsx via openpyxl, xls via xlrd)
# ---------------------------------------------------------------------------

def _read_xlsx_rows(filepath, sheet_name):
    """Read all rows from an xlsx sheet. Returns list of tuples."""
    import openpyxl
    wb = openpyxl.load_workbook(filepath, read_only=True)
    if sheet_name not in wb.sheetnames:
        wb.close()
        return []
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    return rows


def _read_xls_rows(filepath, sheet_name):
    """Read all rows from an xls sheet. Returns list of tuples."""
    import xlrd
    wb = xlrd.open_workbook(filepath)
    if sheet_name not in wb.sheet_names():
        return []
    ws = wb.sheet_by_name(sheet_name)
    rows = []
    for r in range(ws.nrows):
        rows.append(tuple(ws.cell_value(r, c) for c in range(ws.ncols)))
    return rows


def read_rows(filepath, sheet_name):
    """Read rows from either .xlsx or .xls file."""
    if filepath.lower().endswith('.xls'):
        return _read_xls_rows(filepath, sheet_name)
    return _read_xlsx_rows(filepath, sheet_name)


def get_sheet_names(filepath):
    if filepath.lower().endswith('.xls'):
        import xlrd
        wb = xlrd.open_workbook(filepath)
        return wb.sheet_names()
    else:
        import openpyxl
        wb = openpyxl.load_workbook(filepath, read_only=True)
        names = wb.sheetnames
        wb.close()
        return names


# ---------------------------------------------------------------------------
# Section reader using description-based mapping
# ---------------------------------------------------------------------------

def read_section_mapped(filepath, sheet_name):
    """
    Read a PIR section. Returns (descs, codes, records) where:
      descs = row 0 descriptions
      codes = row 1 question codes
      records = list of dicts keyed by question code
    """
    rows = read_rows(filepath, sheet_name)
    if len(rows) < 3:
        return [], [], []

    descs = list(rows[0])
    codes = list(rows[1])

    # Identity columns (first 10)
    identity_keys = codes[:10]

    records = []
    for row in rows[2:]:
        rec = {}
        for i, key in enumerate(identity_keys):
            if key and i < len(row):
                rec[key] = row[i]
        for i in range(10, len(codes)):
            code = codes[i]
            if code and i < len(row):
                rec[code] = row[i]
        records.append(rec)
    return descs, codes, records


def read_program_details(filepath):
    """Read the Program Details sheet (single header row)."""
    rows = read_rows(filepath, "Program Details")
    if len(rows) < 2:
        return []

    headers = rows[0]
    records = []
    for row in rows[1:]:
        rec = {}
        for i, h in enumerate(headers):
            if h and i < len(row):
                rec[str(h).strip()] = row[i]
        records.append(rec)
    return records


# ---------------------------------------------------------------------------
# Merge + map
# ---------------------------------------------------------------------------

def merge_and_map(filepath, details, pir_year, states):
    """
    Read all sections, build description-based code maps, merge by
    (Grant Number, Program Number), and map to headstart_programs columns.
    """
    # Read sections and build code maps
    a_descs, a_codes, sec_a = read_section_mapped(filepath, "Section A")
    b_descs, b_codes, sec_b = read_section_mapped(filepath, "Section B")
    c_descs, c_codes, sec_c = read_section_mapped(filepath, "Section C")
    d_descs, d_codes, sec_d = read_section_mapped(filepath, "Section D")

    a_map = build_code_map(a_descs, a_codes) if sec_a else {}
    b_map = build_staff_code_map(b_descs, b_codes) if sec_b else {}
    c_map = build_health_code_map(c_descs, c_codes) if sec_c else {}
    d_map = build_admin_code_map(d_descs, d_codes) if sec_d else {}

    print(f"    Section A: {len(sec_a)} records, {len(a_map)} fields mapped")
    print(f"    Section B: {len(sec_b)} records, {len(b_map)} fields mapped")
    print(f"    Section C: {len(sec_c)} records, {len(c_map)} fields mapped")
    print(f"    Section D: {len(sec_d)} records, {len(d_map)} fields mapped")

    # Index sections by (grant, program)
    def index_by_key(records):
        idx = {}
        for r in records:
            grant = to_str(r.get("Grant Number") or r.get("Grant_Number"))
            prog = to_str(r.get("Program Number") or r.get("Program_Number"))
            if grant and prog is not None:
                idx[(grant, str(prog).zfill(3))] = r
        return idx

    a_idx = index_by_key(sec_a)
    b_idx = index_by_key(sec_b)
    c_idx = index_by_key(sec_c)
    d_idx = index_by_key(sec_d)

    def get_mapped(record, code_map, field):
        code = code_map.get(field)
        if code and record:
            return record.get(code)
        return None

    rows = []
    for det in details:
        grant = to_str(det.get("Grant Number"))
        prog = to_str(det.get("Program Number"))
        if not grant or prog is None:
            continue
        prog = str(prog).zfill(3)

        state = to_str(det.get("Program State") or det.get("State"))
        if states and state and state.upper() not in [s.upper() for s in states]:
            continue

        key = (grant, prog)
        a = a_idx.get(key, {})
        b = b_idx.get(key, {})
        c = c_idx.get(key, {})
        d = d_idx.get(key, {})

        # Teachers with BA or higher
        ba_teachers = None
        adv = to_int(get_mapped(b, b_map, "_teachers_advanced"))
        bac = to_int(get_mapped(b, b_map, "_teachers_bachelors"))
        if adv is not None or bac is not None:
            ba_teachers = (adv or 0) + (bac or 0)

        row = {
            "grant_number":         grant,
            "program_number":       prog,
            "pir_year":             pir_year,

            "region":               to_str(det.get("Region")),
            "state":                state,
            "program_type":         to_str(det.get("Program Type") or det.get("Type")),
            "grantee_name":         to_str(det.get("Grantee Name") or det.get("Grantee")),
            "program_name":         to_str(det.get("Program Name") or det.get("Program")),
            "agency_type":          to_str(det.get("Program Agency Type")),
            "agency_description":   to_str(det.get("Program Agency Description")),

            "address":              to_str(det.get("Program Address Line 1") or det.get("Address")),
            "city":                 to_str(det.get("Program City") or det.get("City")),
            "zip_code":             to_str(det.get("Program ZIP Code") or det.get("ZIP Code")),
            "phone":                to_str(det.get("Program Main Phone Number") or det.get("Phone")),
            "email":                to_str(det.get("Program Main Email") or det.get("Email")),

            # Section A (description-mapped)
            "funded_enrollment":             to_int(get_mapped(a, a_map, "funded_enrollment")),
            "non_acf_enrollment":            to_int(get_mapped(a, a_map, "non_acf_enrollment")),
            "total_cumulative_enrollment":   to_int(get_mapped(a, a_map, "total_cumulative_enrollment")),
            "total_slots_center_based":      to_int(get_mapped(a, a_map, "total_slots_center_based")),
            "slots_at_child_care_partner":   to_int(get_mapped(a, a_map, "slots_at_child_care_partner")),
            "total_classes":                 to_int(get_mapped(a, a_map, "total_classes")),
            "home_based_slots":              to_int(get_mapped(a, a_map, "home_based_slots")),
            "family_child_care_slots":       to_int(get_mapped(a, a_map, "family_child_care_slots")),

            "children_lt1":                  to_int(get_mapped(a, a_map, "children_lt1")),
            "children_1yr":                  to_int(get_mapped(a, a_map, "children_1yr")),
            "children_2yr":                  to_int(get_mapped(a, a_map, "children_2yr")),
            "children_3yr":                  to_int(get_mapped(a, a_map, "children_3yr")),
            "children_4yr":                  to_int(get_mapped(a, a_map, "children_4yr")),
            "children_5plus":                to_int(get_mapped(a, a_map, "children_5plus")),
            "pregnant_women":                to_int(get_mapped(a, a_map, "pregnant_women")),

            "eligible_income":               to_int(get_mapped(a, a_map, "eligible_income")),
            "eligible_public_assist":        to_int(get_mapped(a, a_map, "eligible_public_assist")),
            "eligible_foster":               to_int(get_mapped(a, a_map, "eligible_foster")),
            "eligible_homeless":             to_int(get_mapped(a, a_map, "eligible_homeless")),

            "children_left_program":         to_int(get_mapped(a, a_map, "children_left_program")),
            "children_end_of_year":          to_int(get_mapped(a, a_map, "children_end_of_year")),

            "dual_language_learners":        to_int(get_mapped(a, a_map, "dual_language_learners")),
            "children_transported":          to_int(get_mapped(a, a_map, "children_transported")),
            "children_with_subsidy":         to_int(get_mapped(a, a_map, "children_with_subsidy")),

            # Section B (description-mapped)
            "total_staff":                   to_int(get_mapped(b, b_map, "total_staff")),
            "total_contracted_staff":        to_int(get_mapped(b, b_map, "total_contracted_staff")),
            "classroom_teachers":            to_int(get_mapped(b, b_map, "classroom_teachers")),
            "assistant_teachers":            to_int(get_mapped(b, b_map, "assistant_teachers")),
            "teachers_ba_or_higher":         ba_teachers,
            "volunteers":                    to_int(get_mapped(b, b_map, "volunteers")),

            # Section C (description-mapped)
            "children_with_insurance_start":    to_int(get_mapped(c, c_map, "children_with_insurance_start")),
            "children_with_insurance_end":      to_int(get_mapped(c, c_map, "children_with_insurance_end")),
            "children_medicaid_start":          to_int(get_mapped(c, c_map, "children_medicaid_start")),
            "children_no_insurance_start":      to_int(get_mapped(c, c_map, "children_no_insurance_start")),
            "children_with_medical_home_start": to_int(get_mapped(c, c_map, "children_with_medical_home_start")),
            "children_at_fqhc_start":           to_int(get_mapped(c, c_map, "children_at_fqhc_start")),

            # Section D (description-mapped)
            "child_care_partners":           to_int(get_mapped(d, d_map, "child_care_partners")),
            "leas_in_service_area":          to_int(get_mapped(d, d_map, "leas_in_service_area")),
        }
        rows.append(row)

    return rows


# ---------------------------------------------------------------------------
# Load one file
# ---------------------------------------------------------------------------

def load_file(filepath, pir_year, states, dry_run=False):
    """Load a single PIR file. Returns number of rows loaded."""
    print(f"\n  [{pir_year}] {os.path.basename(filepath)}")

    sheets = get_sheet_names(filepath)
    if "Program Details" not in sheets:
        print("    SKIP: no Program Details sheet")
        return 0

    details = read_program_details(filepath)
    if not details:
        print("    SKIP: Program Details is empty")
        return 0

    print(f"    {len(details)} programs in Program Details")

    rows = merge_and_map(filepath, details, pir_year, states)
    print(f"    Mapped {len(rows)} program records")

    if not rows:
        return 0

    # Quick type summary
    types = {}
    for r in rows:
        t = r.get("program_type") or "?"
        types[t] = types.get(t, 0) + 1
    print(f"    Types: {dict(sorted(types.items()))}")

    # Check data quality: how many have funded_enrollment?
    has_enrollment = sum(1 for r in rows if r.get("funded_enrollment") is not None)
    print(f"    With funded_enrollment: {has_enrollment}/{len(rows)}")

    if dry_run:
        print("    (dry-run)")
        return 0

    n = db.upsert_rows(
        "headstart_programs", rows,
        unique_cols=["grant_number", "program_number", "pir_year"],
    )
    print(f"    Loaded {n:,} records")
    return n


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    p = argparse.ArgumentParser(description="Load Head Start PIR data from HSES Excel export.")
    p.add_argument("--file", help="Path to a single PIR Excel export (.xlsx or .xls)")
    p.add_argument("--dir", help="Directory of PIR files to batch-load")
    p.add_argument("--year", type=int, default=None, help="Override PIR year")
    p.add_argument("--states", nargs="+", help="Filter to specific states (e.g. GA TX)")
    p.add_argument("--columns-only", action="store_true", help="Print columns and exit")
    p.add_argument("--dry-run", action="store_true", help="Parse but don't write to DB")
    args = p.parse_args()

    if not args.file and not args.dir:
        print("Error: provide --file or --dir")
        return 1

    # Single file: columns-only mode
    if args.file and args.columns_only:
        filepath = args.file
        for name in ["Section A", "Section B", "Section C", "Section D"]:
            rows = read_rows(filepath, name)
            if len(rows) >= 2:
                print(f"\n=== {name} ===")
                descs, codes = rows[0], rows[1]
                for i in range(len(descs)):
                    c = codes[i] if i < len(codes) else ""
                    print(f"  {i:3d}: {str(c):15s} {descs[i]}")
        det_rows = read_rows(filepath, "Program Details")
        if det_rows:
            print(f"\n=== Program Details ===")
            for i, h in enumerate(det_rows[0]):
                print(f"  {i:3d}: {h}")
        return 0

    # Collect files to load
    files = []
    if args.file:
        year = args.year or infer_year(args.file)
        if not year:
            print("Error: could not infer year. Use --year.")
            return 1
        files.append((year, args.file))
    elif args.dir:
        if not os.path.isdir(args.dir):
            print(f"Error: directory not found: {args.dir}")
            return 1
        for fname in sorted(os.listdir(args.dir)):
            if not (fname.lower().endswith('.xlsx') or fname.lower().endswith('.xls')):
                continue
            if not fname.lower().startswith('pir') and not fname.lower().startswith('PIR'):
                continue
            year = args.year or infer_year(fname)
            if year:
                files.append((year, os.path.join(args.dir, fname)))

    if not files:
        print("Error: no PIR files found")
        return 1

    print(f"Head Start PIR Loader")
    print(f"  Files: {len(files)} ({files[0][0]}-{files[-1][0]})")
    if args.states:
        print(f"  States: {', '.join(args.states)}")

    db.init_db()
    run_id = db.log_load_start("headstart_pir")
    total = 0

    try:
        for year, filepath in files:
            n = load_file(filepath, year, args.states, dry_run=args.dry_run)
            total += n
    except Exception as e:
        db.log_load_finish(run_id, rows_loaded=total, error=str(e))
        raise

    db.log_load_finish(run_id, rows_loaded=total)
    print(f"\nDone. Total: {total:,} records loaded across {len(files)} files.")
    return 0


if __name__ == "__main__":
    sys.exit(main())
